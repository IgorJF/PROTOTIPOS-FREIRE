# import os
# import pyautogui
# import time

# procurar = "sim"

# while procurar == "sim":
#     try:
#         img = pyautogui.locateCenterOnScreen("PARABENS/seta.png")
#         pyautogui.click(img.x,img.y)
#         time.sleep(1)
#     except:
#         time.sleep(1)
#         print("não encontrei")

# import openpyxl
# from urllib.parse import quote
# import webbrowser
# from datetime import datetime
# import os 
# import time
# import pyautogui

# workbook = openpyxl.load_workbook('PARABENS/pacientes_aniversario.xlsx')
# sheet = workbook.active

# data_atual = datetime.now().date().strftime("%d-%m")

# for row in sheet.iter_rows(min_row=2, values_only=True): 
#     Nome, Aniversario, Numero = row
    
#     aniversario_mes_dia = Aniversario.date().strftime("%d-%m")
    
#     if aniversario_mes_dia == data_atual:
#         mensagem = f"Olá {Nome}, feliz aniversário! 🎉🎂"

#         Numero = str(Numero)
#         Numero_limpo = ''.join(filter(str.isdigit, Numero))
#         try: 
#             link_whatsapp = f"https://web.whatsapp.com/send?phone={Numero_limpo}&text={quote(mensagem)}"
#             webbrowser.open(link_whatsapp)
#             time.sleep(2) 
#             print("Tentando encontrar a imagem...")
#             img = pyautogui.locateCenterOnScreen("PARABENS/seta3.png")
#             if img:
#                 print("Imagem encontrada!")
#                 pyautogui.click(img.x,img.y)
#                 time.sleep(2)
#                 pyautogui.hotkey('ctrl','w')
#                 time.sleep(2)
#             else:
#                 print("Imagem não encontrada!")
#         except Exception as e:
#            print(f'Erro: {e}')
#            print(f'Não foi possível enviar mensagem para {Nome}')
# workbook.close()

import openpyxl
import time
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.keys import Keys

driver = webdriver.Firefox()

workbook = openpyxl.load_workbook('PARABENS/pacientes_aniversario.xlsx')
sheet = workbook.active

data_atual = datetime.now().date().strftime("%d-%m")

for row in sheet.iter_rows(min_row=2, values_only=True): 
    Nome, Aniversario, Numero = row
    
    aniversario_mes_dia = Aniversario.date().strftime("%d-%m")
    
    if aniversario_mes_dia == data_atual:
        audio_path = "PARABENS/teste.mp3" 
        Numero = str(Numero)
        Numero_limpo = ''.join(filter(str.isdigit, Numero))
        try: 
            link_whatsapp = f"https://web.whatsapp.com/send?phone={Numero_limpo}"
            driver.get(link_whatsapp)
            
            time.sleep(30) 
            input_box = driver.find_element_by_xpath('//*[@id="main"]/footer/div[1]/div[2]/div/div[2]')
            
            input_box.send_keys(audio_path)
            time.sleep(2)
            
            input_box.send_keys(Keys.RETURN)
            time.sleep(2)
        except:
            print(f'Não foi possível enviar mensagem para {Nome}')
            
driver.quit()
workbook.close()
